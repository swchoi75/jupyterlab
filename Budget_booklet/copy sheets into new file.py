from openpyxl import load_workbook, Workbook
import os

# Define the directory where the Excel files are located
directory = "path_to_directory_containing_files"

# Create a new workbook
new_workbook = Workbook()
new_workbook.remove(new_workbook.active)

# List Excel files in the specified directory
excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]

for file_name in excel_files:
    # Extract the sheet name from the file name (excluding the extension)
    sheet_name = os.path.splitext(file_name)[0]

    # Load the source workbook
    source_workbook = load_workbook(
        os.path.join(directory, file_name), data_only=True)

    if sheet_name in source_workbook.sheetnames:
        source_sheet = source_workbook[sheet_name]
        new_sheet = new_workbook.create_sheet(title=sheet_name)
        for row in source_sheet.iter_rows(values_only=True):
            new_sheet.append(row)

# Save the new workbook with copied sheets
new_workbook.save("new_combined_file.xlsx")
