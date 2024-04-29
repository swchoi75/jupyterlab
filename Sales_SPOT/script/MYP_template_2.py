import csv
import openpyxl as op
from openpyxl.worksheet.formula import ArrayFormula
from pathlib import Path


# Path
path = Path(__file__).parent.parent


# Functions
def add_text(worksheet):
    # ws["C6"] = "Instruction / comments"
    worksheet.cell(row=6, column=3).value = "Instruction / comments"


def add_items(worksheet, list_of_data):
    # Start from the cell. Rows and columns are zero indexed.
    row = 1
    col = 4

    # Write the data from the CSV file, handling empty lines
    for data in list_of_data:
        # Check if the row is not empty before accessing data[0]
        if data:
            worksheet.cell(row, col).value = data[1]  # items
        row += 1


def add_poc(worksheet, list_of_data):
    # Start from the cell. Rows and columns are zero indexed.
    row = 2
    col = 5

    for data in list_of_data:
        worksheet.cell(row, col).value = data[0]  # Division
        col += 9

    # Start from the cell. Rows and columns are zero indexed.
    row = 3
    col = 5

    for data in list_of_data:
        worksheet.cell(row, col).value = data[1]  # BU
        col += 9

    # Start from the cell. Rows and columns are zero indexed.
    row = 4
    col = 5

    for data in list_of_data:
        worksheet.cell(row, col).value = data[2]  # Productlines
        col += 9


def add_years(worksheet, list_of_data):
    # Start from the cell. Rows and columns are zero indexed.
    rows = [6, 48, 77, 105, 123, 135, 144]

    for row in rows:
        col = 5
        for data in list_of_data:
            worksheet.cell(row, col).value = data
            col += 1


def add_zero(worksheet, list_of_data):
    # Start from the cell. Rows and columns are zero indexed.
    rows = (
        list(range(8, 45 + 1))
        + list(range(106, 113 + 1))
        + list(range(124, 130 + 1))
        + list(range(136, 138 + 1))
        + list(range(145, 147 + 1))
    )

    for row in rows:
        col = 5
        for data in list_of_data:
            worksheet.cell(row, col).value = data
            col += 1


def add_formula_sg(worksheet):
    # Dynamic array
    worksheet["E79"] = ArrayFormula("E79:L79", "=E50:L50-E51:L51-E52:L52")
    # worksheet["E79"] = ArrayFormula(
    #     "E79:L79", "(E10:L10+E12:L12+E13:L13+E16:L16+E22:L22+E30:L30+E31:L31)*-1"
    # )


def main():
    # Filenames
    input_items = path / "data" / "items.csv"
    input_pl = path / "data" / "PLs.csv"
    input_file = path / "data" / "MYP consolidation template.xlsx"
    output_file = path / "output" / "MYP_template_op.xlsx"
    sheet_names = ["Korea", "India", "Japan", "Thailand"]

    # Read data
    with open(input_items, "r", encoding="utf-8") as f:
        csv_reader = csv.reader(f)
        items = list(csv_reader)

    with open(input_pl, "r", encoding="utf-8") as f:
        csv_reader = csv.reader(f)
        data = list(csv_reader)

    # Year data
    fy = [2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, ""] * 27  # no. of section

    # Zero values
    zero = [0, 0, 0, 0, 0, 0, 0, 0, ""] * 27  # no. of section

    # Create an Excel workbook
    wb = op.Workbook()
    ws = wb.active

    add_text(ws)
    add_items(ws, items)
    add_poc(ws, data)
    add_years(ws, fy)
    add_zero(ws, zero)
    # add_formula_vt(ws)
    add_formula_sg(ws)

    wb.save(output_file)

    print("A file is created")

    # Read from an Excel workbook
    # wb = op.load_workbook(input_file)
    # ws_list = wb.sheetnames
    # print(ws_list)


if __name__ == "__main__":
    main()
