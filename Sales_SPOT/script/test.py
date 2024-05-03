from openpyxl import load_workbook

# Load the workbook
wb = load_workbook("path/to/your/original/file.xlsx")

# List all sheet names
all_sheets = wb.sheetnames

# Sheet to keep
sheet_to_keep = "Sheet1"

# Remove other sheets
for sheet in all_sheets:
    if sheet != sheet_to_keep:
        del wb[sheet]

# Save the modified workbook to a new file
wb.save("path/to/your/new/file.xlsx")

print("Workbook has been saved with only the desired sheet.")
